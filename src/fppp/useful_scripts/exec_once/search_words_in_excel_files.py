import os
import openpyxl

DEFAULT_WORDS = ['nick', 'nickname', 'ник', 'username', 'usermail', 'email', 'e-mail', 'useremail', 'emailaddress',
                 'майл', 'е-мейл', 'е-мэйл', 'e-mailaddress', 'tel', 'phone', 'mobile', 'mobilenumber', 'mobilephone',
                 'telephone', 'phonenumber', 'телефон', 'telefono', 'тел', 'userfname', 'fname', 'firstname', 'имя',
                 'fullname', 'фамилияиимя', 'имяифамилия', 'фио', 'фио', 'pib', 'name', 'lname', 'userlname',
                 'lastname', 'фамилия', 'фам', 'familia', 'lastname', 'surname', 'password', 'passwd', 'passwordhash',
                 'passhash', 'passwordsalt', 'passsalt', 'address', 'address2', 'address3', 'streetaddress', 'адрес',
                 'адреcс', 'address1', 'homeaddress', 'adres', 'city', 'town', 'city/town', 'город', 'homecity',
                 'addressaddress', 'регион', 'область', 'обл', 'homestate', 'addressstate', 'state', 'country',
                 'страна', 'homecountry', 'addresscountry', 'zipcode', 'postcode', 'addresszip', 'postalcode', 'zip',
                 'psprt', 'passport', 'pasport', 'паспорт', 'pasportnumber', 'passportnumber', 'dateofbirth',
                 'datebirth']


def search_words_in_files(folder_path, words_list=DEFAULT_WORDS):
    # Перебираем все файлы в папке
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            matched_words = []
            # Получаем полный путь до файла
            file_path = os.path.join(root, file_name)
            # Ищем нужные слова в файле Excel
            if file_name.endswith('.xlsx') or file_name.endswith('.xlsm'):
                try:
                    workbook = openpyxl.load_workbook(file_path)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                # Проверяем, содержит ли ячейка нужное слово
                                for word in words_list:
                                    if word in str(cell.value):
                                        matched_words.append(word)
                finally:
                    pass
            # Ищем нужные слова в остальных файлах
            else:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    for line in file:
                        # Проверяем, содержит ли строка нужное слово
                        for word in words_list:
                            if word in line:
                                matched_words.append(word)
            if len(matched_words) > 1:
                print(
                    f'Слова "{matched_words}" найдены в файле {file_path}')
